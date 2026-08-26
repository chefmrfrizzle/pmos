from __future__ import annotations
import csv, hashlib, json, os, re, zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable, Sequence
from urllib.parse import urlparse
from openpyxl import load_workbook
from sqlalchemy import select
from .db import Claim, Contact, Entity, Evidence, ImportBatch, RawImportRow, ResolutionDecision
from .entity_resolution import MatchState, canonicalize_name, resolve

IDENTITY_HEADERS={"name","contact name","contact","investor name","firm","organization","organisation","company","institution"}
CONTACT_HEADERS=("contact name","full name","person","contact","name")
FIRM_HEADERS=("investor name","fund name","institution","organization","organisation","company","firm")
EMAIL_HEADERS=("email address","e-mail","email"); PHONE_HEADERS=("phone number","telephone","mobile","phone")
TITLE_HEADERS=("position / role","job title","position","title","role")
COUNTRY_HEADERS=("country / jurisdiction","hq country","jurisdiction","country"); CITY_HEADERS=("hq city","headquarters","location","city")
TYPE_HEADERS=("institution type","investor type","category","segment","type"); URL_HEADERS=("official url","company website","website","url")
MANDATE_HEADERS=("investment focus","mandate","strategy","description","notes")
ROLE_INBOXES={"info","contact","office","admin","hello","team","sales","support","enquiries","inquiries","press","media"}
MAX_FILE_BYTES=max(1024*1024,min(int(os.getenv("PMOS_IMPORT_MAX_FILE_BYTES","100000000")),500000000))
MAX_ROWS=max(1000,min(int(os.getenv("PMOS_IMPORT_MAX_ROWS","100000")),500000))
MAX_COLUMNS=max(10,min(int(os.getenv("PMOS_IMPORT_MAX_COLUMNS","500")),2000))
MAX_CELL_CHARS=max(1000,min(int(os.getenv("PMOS_IMPORT_MAX_CELL_CHARS","100000")),1000000))
MAX_TOTAL_CELLS=max(10000,min(int(os.getenv("PMOS_IMPORT_MAX_TOTAL_CELLS","5000000")),20000000))
MAX_XLSX_UNCOMPRESSED=max(10_000_000,min(int(os.getenv("PMOS_IMPORT_MAX_XLSX_UNCOMPRESSED","500000000")),2_000_000_000))

class ImportSafetyError(ValueError):pass

def preflight_import(path:Path,expected_suffix:str)->None:
    if path.suffix.casefold()!=expected_suffix or not path.is_file() or path.is_symlink():raise ImportSafetyError("unsupported or unsafe import file")
    size=path.stat().st_size
    if size<=0 or size>MAX_FILE_BYTES:raise ImportSafetyError("import file size is outside configured limits")
    if expected_suffix==".xlsx":
        try:
            with zipfile.ZipFile(path) as archive:
                members=archive.infolist()
                if len(members)>5000:raise ImportSafetyError("workbook has too many archive members")
                compressed=sum(max(x.compress_size,1) for x in members);uncompressed=sum(x.file_size for x in members)
                if uncompressed>MAX_XLSX_UNCOMPRESSED or uncompressed/compressed>100:raise ImportSafetyError("workbook archive expansion exceeds configured limits")
                if any(x.file_size>MAX_XLSX_UNCOMPRESSED//2 or x.flag_bits&1 for x in members):raise ImportSafetyError("workbook contains an unsafe archive member")
        except zipfile.BadZipFile as exc:raise ImportSafetyError("invalid XLSX archive") from exc

def _bounded_rows(rows):
    output=[];cells=0
    for index,row in rows:
        if len(output)>=MAX_ROWS:raise ImportSafetyError("import row limit exceeded")
        if len(row)>MAX_COLUMNS:raise ImportSafetyError("import column limit exceeded")
        cells+=len(row)
        if cells>MAX_TOTAL_CELLS:raise ImportSafetyError("import cell limit exceeded")
        values=tuple(row)
        if any(len(str(value))>MAX_CELL_CHARS for value in values if value is not None):raise ImportSafetyError("import cell size limit exceeded")
        output.append((index,values))
    return output

def _val(value)->str:
    if value is None:return ""
    return str(value).strip()
def _header(value)->str:return re.sub(r"\s+"," ",_val(value).lower()).strip(" :")
def _json(value)->str:return json.dumps(value,ensure_ascii=False,sort_keys=True,separators=(",",":"),default=str)
def _hash(value)->str:return hashlib.sha256(_json(value).encode()).hexdigest()
def file_sha256(path:Path)->str:
    digest=hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda:handle.read(1024*1024),b""):digest.update(block)
    return digest.hexdigest()
def detect_header(rows:Sequence[Sequence[object]])->int|None:
    best=None
    for index,row in enumerate(rows[:25]):
        labels={_header(v) for v in row if _val(v)}; identity=len(labels&IDENTITY_HEADERS); score=(identity>0,identity,len(labels),-index)
        if identity and (best is None or score>best[0]):best=(score,index)
    return None if best is None else best[1]
def _first(row:dict[str,str],aliases)->str:return next((row[key] for key in aliases if row.get(key)),"")
def _source(path:Path,row_hash:str)->str:return f"private-import://{path.name}/{row_hash}"

def _candidate_entity(session,name,country,url):
    candidates=session.scalars(select(Entity).where(Entity.canonical_name==canonicalize_name(name)).limit(20)).all()
    incoming_domain=urlparse(url).netloc.lower().removeprefix("www.")
    exact=[x for x in candidates if country and x.country and x.country.casefold()==country.casefold() and incoming_domain and x.official_url and urlparse(x.official_url).netloc.lower().removeprefix("www.")==incoming_domain]
    if len(exact)==1:return exact[0],MatchState.EXACT,.98,("same normalized name","same jurisdiction","same official domain")
    if len(candidates)==1:
        result=resolve({"name":name,"url":url},{"name":candidates[0].name,"url":candidates[0].official_url})
        return candidates[0],result.state,result.confidence,result.reasons
    if candidates:return candidates[0],MatchState.REVIEW,.45,("multiple normalized-name candidates",)
    return None,MatchState.REVIEW,0,("no existing candidate",)
def _candidate_contact(session,name,email):
    if email:
        matches=session.scalars(select(Contact).where(Contact.email.ilike(email)).limit(3)).all()
        local=email.split("@",1)[0].casefold()
        if local in ROLE_INBOXES and matches:return matches[0],MatchState.REVIEW,.35,("shared or role inbox cannot identify a person",)
        if len(matches)==1:
            if canonicalize_name(name)==canonicalize_name(matches[0].name):return matches[0],MatchState.EXACT,.99,("same normalized email","compatible normalized name")
            return matches[0],MatchState.CONFLICT,.2,("same email with conflicting person name",)
        if len(matches)>1:return matches[0],MatchState.CONFLICT,.25,("email belongs to multiple records",)
    if name:
        matches=session.scalars(select(Contact).where(Contact.name.ilike(name)).limit(3)).all()
        if len(matches)==1:return matches[0],MatchState.PROBABLE,.86,("same display name","no unique email evidence")
        if len(matches)>1:return matches[0],MatchState.REVIEW,.4,("multiple name candidates",)
    return None,MatchState.REVIEW,0,("no existing candidate",)
def _decision(session,raw,entity,contact,state,confidence,reasons):
    session.add(ResolutionDecision(raw_row_id=raw.id,candidate_entity_id=getattr(entity,"id",None),candidate_contact_id=getattr(contact,"id",None),state=state.value,confidence=confidence,reasons_json=_json(list(reasons)),automatic=True))
def _claim(session,entity,field,value,source,row_hash,confidence):
    if value:session.add(Claim(entity_id=entity.id,field=field,value=value,source_url=source,source_type="private_import",confidence=confidence,verification_status="CANDIDATE",extractor="deterministic_private_import_v2",evidence_hash=row_hash))

def _materialize(session,path,raw,row):
    firm=_first(row,FIRM_HEADERS); name=_first(row,CONTACT_HEADERS); email=_first(row,EMAIL_HEADERS).lower()
    phone=_first(row,PHONE_HEADERS); title=_first(row,TITLE_HEADERS); country=_first(row,COUNTRY_HEADERS); city=_first(row,CITY_HEADERS)
    entity_type=_first(row,TYPE_HEADERS); url=_first(row,URL_HEADERS); mandate=_first(row,MANDATE_HEADERS)
    if not firm and not name:return "support_row"
    entity=None
    if firm:
        candidate,state,confidence,reasons=_candidate_entity(session,firm,country,url)
        if state is MatchState.EXACT:entity=candidate
        else:
            entity=Entity(name=firm,canonical_name=canonicalize_name(firm),universe="imported_private",entity_type=entity_type or None,country=country or None,city=city or None,official_url=url or None,mandate=mandate or None,verification_status="CANDIDATE")
            session.add(entity);session.flush()
        raw.entity_id=entity.id; source=_source(path,raw.row_hash)
        session.add(Evidence(entity_id=entity.id,source_url=source,source_type="private_import",content_hash=raw.row_hash,title=f"Private import row · {path.name}",confidence=.45))
        for field,value in (("name",firm),("entity_type",entity_type),("jurisdiction",country),("city",city),("official_url",url),("mandate",mandate)):_claim(session,entity,field,value,source,raw.row_hash,.55 if field=="name" else .45)
        _decision(session,raw,candidate,None,state,confidence,reasons)
    if name:
        candidate,state,confidence,reasons=_candidate_contact(session,name,email)
        if state is MatchState.EXACT:
            contact=candidate
            if entity and contact.entity_id is None:contact.entity_id=entity.id
        else:
            contact=Contact(entity_id=getattr(entity,"id",None),name=name,title=title or None,email=email or None,phone=phone or None,source=f"private-import://{path.name}",relationship_stage="not_contacted")
            session.add(contact);session.flush()
        raw.contact_id=contact.id;_decision(session,raw,None,candidate,state,confidence,reasons)
    return "imported"

def _begin(session,path):
    digest=file_sha256(path); prior=session.scalar(select(ImportBatch).where(ImportBatch.source_sha256==digest).order_by(ImportBatch.id))
    if prior:return prior,False
    batch=ImportBatch(source_file=path.name,source_sha256=digest);session.add(batch);session.flush();return batch,True
def _record_rows(session,path,batch,sheet,rows:Iterable[tuple[int,Sequence[object]]],header_index):
    buffered=list(rows); header=[]
    if header_index is not None:header=[_header(v) or f"column_{i+1}" for i,v in enumerate(buffered[header_index][1])]
    count=0
    for ordinal,(physical,row_values) in enumerate(buffered):
        values=[_val(v) for v in row_values]
        if not any(values):continue
        count+=1;batch.rows_seen+=1;normalized={}
        if header_index is not None and ordinal>header_index:normalized={header[i] if i<len(header) else f"column_{i+1}":value for i,value in enumerate(values) if value}
        disposition="header" if ordinal==header_index else "preamble" if header_index is not None and ordinal<header_index else "support_row" if header_index is None else "pending"
        row_hash=_hash({"sheet":sheet,"row":physical,"values":values})
        raw=RawImportRow(batch_id=batch.id,source_file=path.name,sheet_name=sheet,source_row_number=physical,row_hash=row_hash,original_row_json=_json(values),normalized_row_json=_json(normalized),disposition=disposition)
        session.add(raw);session.flush()
        if disposition=="pending":disposition=_materialize(session,path,raw,normalized);raw.disposition=disposition
        if disposition=="imported":batch.rows_imported+=1
        elif disposition=="requires_review":batch.rows_review+=1
        else:batch.rows_support+=1
    return count
def import_csv(session,path:Path)->int:
    preflight_import(path,".csv")
    batch,is_new=_begin(session,path)
    if not is_new:return 0
    try:
        csv.field_size_limit(MAX_CELL_CHARS)
        with path.open("r",encoding="utf-8-sig",errors="replace",newline="") as handle:rows=_bounded_rows((i,row) for i,row in enumerate(csv.reader(handle),1))
        count=_record_rows(session,path,batch,"CSV",rows,0 if rows else None);batch.status="completed";batch.completed_at=datetime.now(timezone.utc);return count
    except Exception as exc:batch.status="failed";batch.error=f"{type(exc).__name__}: {exc}"[:2000];raise
def import_xlsx(session,path:Path)->int:
    preflight_import(path,".xlsx")
    batch,is_new=_begin(session,path)
    if not is_new:return 0
    count=0
    try:
        workbook=load_workbook(path,read_only=True,data_only=False)
        for sheet in workbook.worksheets:
            rows=_bounded_rows((i,tuple(row)) for i,row in enumerate(sheet.iter_rows(values_only=True),1))
            count+=_record_rows(session,path,batch,sheet.title,rows,detect_header([row for _,row in rows]))
        workbook.close();batch.status="completed";batch.completed_at=datetime.now(timezone.utc);return count
    except Exception as exc:batch.status="failed";batch.error=f"{type(exc).__name__}: {exc}"[:2000];raise
